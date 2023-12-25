from openpyxl import Workbook, load_workbook
import tabula
import pandas as pd
import numpy as np

'''擷取pdf'''
# 設定欄位名稱
columns = ["交易日期", "帳務日期", "交易代號", "交易時間", "交易分行 交易櫃員", "摘要", "支出", "存入", "餘額", "轉出入帳號", "合作機構會員編號", "金資序號", "票號", "備註", "註記"]

df = tabula.read_pdf("客戶備註很多.pdf", area=[120, 5, 800, 1200], pages="all")
# 初始化合併的 DataFrame
all_df = pd.DataFrame()

# 逐頁處理
for page_df in df:
    # 如果有表格，轉換為 DataFrame
    if not page_df.empty:
        # 從第二行開始合併
        page_df = page_df.iloc[1:]

        # 將 DataFrame 合併到結果中
        all_df = pd.concat([all_df, page_df], ignore_index=True)

# 前15個欄位
result_df = all_df.iloc[:, :15]
df = result_df.fillna(0)
df.columns = columns


'''寫入excel'''
# 讀取excel檔
wb = load_workbook("關鍵字統計.xlsx")
ws = wb.active

#將 "支出、存入" 列中的逗號刪除並轉換為浮點數
df["支出"] = df["支出"].str.replace(",", "").astype(float)
df["存入"] = df["存入"].str.replace(",", "").astype(float)

# 定義類別與相對應的關鍵字
categories = {
    "放貸/會錢": ["日會","會錢","貸款","利息","借款","還款","結清","本金"],
    "虛擬貨幣": ["英屬維京群島商","現代財富科技有","凱基商業銀行受","幣","USTD","binance","BINGX"],
    "地下匯兌/代儲": ["兌換","換幣","草","RMB","rmb","人民幣","USD","幫幫寶","支付寶"],
    "遊戲/代儲": ["遊戲","代儲","星城","王者","天堂","天2","鑽","儲值","TIKTOK","抖音","微信"],
    "電子菸": ["電子菸","彈","煙","菸","油","IQOS","iqos"],
    "網拍/代購": ["鞋","包","貨款","香奈兒","香水","衣服","手錶","皮帶","代購","Gucci","LV","Prada","精品"],
    "股票/代操": ["股票","股款","配股","配息","認股","交割"],
    "台彩": ["台彩","刮","發財金","營收","兌獎金額","銷售佣金","銷售額度"],
    "運彩/博弈": ["運彩","牌","分析費","世足","贏","發財金","投資","紅利","重注","娛樂","賽事","MITRADE"]
}

#初始欄位
start_row, start_column = "F",7

def reset_start_cell():
    global start_row, start_column
    start_row ="F"
    start_column += 5

# 使用迴圈處理每個關鍵字
for category,keywords in categories.items():
    for keyword in keywords:
        # 將新的列號和行號重新組合成字符串
        start_cell = f"{start_row}{start_column}"
        # 選擇符合關鍵字條件的資料行
        selected_data = df[df["備註"].str.contains(keyword, na=False)]
    
        # 計算支出和存入的總和
        total_expense = selected_data["支出"].sum()
        total_income = selected_data["存入"].sum()
        
        # 初始化支出和存入的次數
        expense_occurrences = 0        
        income_occurrences = 0
        # 計算出現的次數
        expense_occurrences += selected_data["支出"].count() if total_expense > 0 else 0
        income_occurrences += selected_data["存入"].count() if total_income > 0 else 0

        ws[start_cell].value = keyword
        ws[start_cell].offset(1, 0).value = income_occurrences #"存入次數"
        ws[start_cell].offset(2, 0).value = total_income #"存入金額"
        ws[start_cell].offset(3, 0).value = expense_occurrences #"支出次數"
        ws[start_cell].offset(4, 0).value = total_expense #"支出金額"
        # 計算下一個類別的起始欄位
        start_row = chr(ord(start_row) + 1)

    # 計算下一個類別的起始欄位
    reset_start_cell()

# 保存 Excel 文件
wb.save("客戶備註很多關鍵字統計.xlsx")        
